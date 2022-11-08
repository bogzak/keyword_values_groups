# pip install openpyxl

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

file_phrases = load_workbook("phrases_dom_value.xlsx") # название файла xlsx
sheet_phrases = file_phrases["Sheet1"] # название листа

row = 2
row_start = 2
for rows in sheet_phrases.iter_rows(max_col=5, min_row=2):
    if sheet_phrases.cell(column=2, row=row).value == sheet_phrases.cell(column=2, row=row+1).value:
        pass
    elif sheet_phrases.cell(column=2, row=row).value == "None":
        break
    else:
        sheet_phrases.insert_rows(row+1, 1)
        sheet_phrases.cell(row=row+1, column=1, value="Итог").fill = PatternFill("solid", fgColor="5eba7d")
        sheet_phrases.cell(row=row+1, column=2, value=sheet_phrases.cell(row=row, column=2).value).fill = PatternFill("solid", fgColor="5eba7d")
        sheet_phrases.cell(row=row+1, column=3, value=f"=AVERAGE(C{row_start}:C{row})").fill = PatternFill("solid", fgColor="5eba7d")
        sheet_phrases.cell(row=row+1, column=4, value=f"=SUM(D{row_start}:D{row})").fill = PatternFill("solid", fgColor="5eba7d")
        sheet_phrases.cell(row=row+1, column=5, value=f"=SUM(E{row_start}:E{row})").fill = PatternFill("solid", fgColor="5eba7d")
        row += 1
        row_start = row+1
    row += 1

file_phrases.save("phrases_dom_value.xlsx") # сохранить файл